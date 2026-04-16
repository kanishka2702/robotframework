import openpyxl

class ExcelReader:

    def read_credentials_from_excel(self, file_path, sheet_name="Login"):
        """
        Reads username and password from Excel.
        Expects headers in row 1: username | password
        Reads data from row 2
        """
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        username = sheet.cell(row=2, column=1).value
        password = sheet.cell(row=2, column=2).value

        workbook.close()
        return username, password


def get_excel_data_list(self, file_path, excel_sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[excel_sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]